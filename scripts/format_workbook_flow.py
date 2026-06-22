from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

FILE = Path("reports/mlb-lab-v5-matchup-engine.xlsx")

wb = load_workbook(FILE)

# remove old dashboard if exists
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]

dash = wb.create_sheet("Dashboard", 0)

dark = "1F4E78"
blue = "D9EAF7"
green = "D9EAD3"
white = "FFFFFF"
thin = Side(style="thin", color="C9D3DF")

dash["A1"] = "MLB-LAB V5 Matchup Engine"
dash["A1"].font = Font(size=18, bold=True, color=white)
dash["A1"].fill = PatternFill("solid", fgColor=dark)

dash["A3"] = "Use this sheet like a slate control panel. Click a game tab, review pitch mix, hitter pools, bullpen fatigue, and final dissection."
dash["A3"].font = Font(size=11, italic=True)

headers = ["Game #", "Game Tab", "Open", "What To Inspect"]
for c, h in enumerate(headers, 1):
    cell = dash.cell(5, c, h)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=dark)
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

row = 6
for ws in wb.worksheets:
    if ws.title in ["Dashboard", "Slate", "Master Matchups"]:
        continue
    dash.cell(row, 1, row - 5)
    dash.cell(row, 2, ws.title)
    dash.cell(row, 3, f'=HYPERLINK("#\'{ws.title}\'!A1","Open Game")')
    dash.cell(row, 4, "Pitch arsenal → hitter vs pitch mix → hitter pool → bullpen → final read")
    for c in range(1, 5):
        cell = dash.cell(row, c)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    row += 1

dash.column_dimensions["A"].width = 10
dash.column_dimensions["B"].width = 24
dash.column_dimensions["C"].width = 18
dash.column_dimensions["D"].width = 70
dash.freeze_panes = "A6"
dash.sheet_view.showGridLines = False

# format every game sheet for cleaner flow
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 85

    for col in range(1, min(ws.max_column, 14) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row_cells in ws.iter_rows():
        values = [str(c.value or "") for c in row_cells]
        text = " ".join(values).lower()
        if any(key in text for key in ["pitch arsenal", "hitter pool", "bullpen", "last 5", "final game dissection"]):
            for cell in row_cells:
                cell.fill = PatternFill("solid", fgColor=dark)
                cell.font = Font(bold=True, color=white)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.save(FILE)
print(f"Formatted workbook flow: {FILE}")
