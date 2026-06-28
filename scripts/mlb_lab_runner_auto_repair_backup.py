from datetime import date, timedelta
from pathlib import Path
import requests
import pandas as pd
from pybaseball import statcast, playerid_reverse_lookup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

TODAY = date.today()
START = (TODAY - timedelta(days=120)).isoformat()
END = TODAY.isoformat()

REPORT = Path("reports/mlb-lab-v5-matchup-engine.md")
WORKBOOK = Path("reports/mlb-lab-v5-matchup-engine.xlsx")
CACHE = Path("data/cache/statcast_cache.pkl")

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri")
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FILL = PatternFill("solid", start_color="2E5395", end_color="2E5395")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SUBHEAD_FONT = Font(name="Calibri", bold=True, size=10, color="333333")
GAME_FILL_A = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
GAME_FILL_B = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

# Stats where HIGH = favorable for hitter (green high, red low)
GOOD_HIGH_STATS = {"AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%"}
# Stats where HIGH = favorable for pitcher / bad for hitter (red high, green low)
GOOD_LOW_STATS = {"Whiff%", "K%", "K Event Rate", "HR Event Rate"}

TEAM_CODES = {
    "Arizona Diamondbacks": "AZ", "Athletics": "ATH", "Atlanta Braves": "ATL",
    "Baltimore Orioles": "BAL", "Boston Red Sox": "BOS", "Chicago Cubs": "CHC",
    "Chicago White Sox": "CWS", "Cincinnati Reds": "CIN", "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL", "Detroit Tigers": "DET", "Houston Astros": "HOU",
    "Kansas City Royals": "KC", "Los Angeles Angels": "LAA", "Los Angeles Dodgers": "LAD",
    "Miami Marlins": "MIA", "Milwaukee Brewers": "MIL", "Minnesota Twins": "MIN",
    "New York Mets": "NYM", "New York Yankees": "NYY", "Philadelphia Phillies": "PHI",
    "Pittsburgh Pirates": "PIT", "San Diego Padres": "SD", "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA", "St. Louis Cardinals": "STL", "Tampa Bay Rays": "TB",
    "Texas Rangers": "TEX", "Toronto Blue Jays": "TOR", "Washington Nationals": "WSH",
}

def mlb_json(url):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return r.json()

def fmt(x, n=3):
    try:
        if x is None or pd.isna(x):
            return "—"
        return f"{float(x):.{n}f}"
    except Exception:
        return "—"

def pct(x):
    try:
        if x is None or pd.isna(x):
            return "—"
        return f"{float(x) * 100:.1f}%"
    except Exception:
        return "—"

def md_table(headers, rows):
    out = "| " + " | ".join(headers) + " |\n"
    out += "| " + " | ".join(["---"] * len(headers)) + " |\n"
    if not rows:
        rows = [["No data"] + ["—"] * (len(headers) - 1)]
    for row in rows:
        out += "| " + " | ".join(str(x) for x in row) + " |\n"
    return out

def get_schedule():
    url = f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={TODAY}&hydrate=probablePitcher,venue,team"
    data = mlb_json(url)
    games = []

    for d in data.get("dates", []):
        for g in d.get("games", []):
            away = g["teams"]["away"]["team"]
            home = g["teams"]["home"]["team"]
            away_sp = g["teams"]["away"].get("probablePitcher", {})
            home_sp = g["teams"]["home"].get("probablePitcher", {})

            games.append({
                "game": f"{away['name']} @ {home['name']}",
                "away_team": away["name"],
                "home_team": home["name"],
                "away_team_id": away["id"],
                "home_team_id": home["id"],
                "away_code": TEAM_CODES.get(away["name"], ""),
                "home_code": TEAM_CODES.get(home["name"], ""),
                "park": g["venue"]["name"],
                "time": g.get("gameDate", "—"),
                "away_sp": away_sp.get("fullName", "TBD"),
                "home_sp": home_sp.get("fullName", "TBD"),
                "away_sp_id": away_sp.get("id"),
                "home_sp_id": home_sp.get("id"),
            })

    return games

def load_statcast():
    CACHE.parent.mkdir(parents=True, exist_ok=True)

    try:
        print(f"Pulling Statcast {START} to {END}")
        sc = statcast(start_dt=START, end_dt=END)
        if not sc.empty:
            sc.to_pickle(CACHE)
            return prep_statcast(sc)
    except Exception as e:
        print(f"Statcast pull failed: {e}")

    if CACHE.exists():
        print("Using cached Statcast.")
        return prep_statcast(pd.read_pickle(CACHE))

    return pd.DataFrame()

def prep_statcast(sc):
    sc = sc.copy()

    if "inning_topbot" in sc.columns:
        sc["batter_team"] = sc.apply(
            lambda r: r.get("away_team") if r.get("inning_topbot") == "Top" else r.get("home_team"),
            axis=1
        )

    if "batter" in sc.columns:
        try:
            ids = sorted(sc["batter"].dropna().astype(int).unique().tolist())
            lookup = playerid_reverse_lookup(ids, key_type="mlbam")
            lookup["batter_name"] = lookup["name_first"] + " " + lookup["name_last"]
            sc["batter_name"] = sc["batter"].map(dict(zip(lookup["key_mlbam"], lookup["batter_name"])))
        except Exception as e:
            print(f"Batter lookup failed: {e}")
            sc["batter_name"] = ""

    return sc

def recent_completed_games(team_id, days_back=4):
    if not team_id:
        return []

    start = (TODAY - timedelta(days=days_back)).isoformat()
    end = (TODAY - timedelta(days=1)).isoformat()
    url = (f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&teamId={team_id}"
           f"&startDate={start}&endDate={end}&gameType=R")

    try:
        data = mlb_json(url)
    except Exception as e:
        print(f"Schedule lookup failed for team {team_id}: {e}")
        return []

    game_pks = []
    for d in data.get("dates", []):
        for g in d.get("games", []):
            if g.get("status", {}).get("abstractGameState") == "Final":
                game_pks.append((d.get("date"), g.get("gamePk")))

    return sorted(game_pks)

def bullpen_usage(team_id, days_back=4):
    games = recent_completed_games(team_id, days_back)

    if not games:
        return [], set()

    rows = []
    appearance_count = {}
    yesterday = (TODAY - timedelta(days=1)).isoformat()

    for game_date, game_pk in games:
        try:
            box = mlb_json(f"https://statsapi.mlb.com/api/v1/game/{game_pk}/boxscore")
        except Exception as e:
            print(f"Boxscore lookup failed for game {game_pk}: {e}")
            continue

        side = None
        for s in ("home", "away"):
            if box.get("teams", {}).get(s, {}).get("team", {}).get("id") == team_id:
                side = s
                break

        if not side:
            continue

        team_box = box["teams"][side]
        pitcher_ids = team_box.get("pitchers", [])
        players = team_box.get("players", {})

        # First pitcher listed for a team in a completed game is the starter;
        # everyone after is bullpen usage.
        for idx, pid in enumerate(pitcher_ids):
            stats = players.get(f"ID{pid}", {}).get("stats", {}).get("pitching", {})

            if not stats or idx == 0:
                continue

            name = players.get(f"ID{pid}", {}).get("person", {}).get("fullName", "Unknown")
            ip = stats.get("inningsPitched", "0.0")
            pitches = stats.get("numberOfPitches", 0)

            rows.append([name, game_date, ip, pitches])
            appearance_count[name] = appearance_count.get(name, 0) + 1

    tired = {name for name, count in appearance_count.items() if count >= 3}
    tired |= {r[0] for r in rows if r[1] == yesterday}

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows, tired

def bullpen_context(team_id, team_name, days_back=4):
    headers = ["Reliever", "Date", "IP", "Pitches"]
    rows, tired = bullpen_usage(team_id, days_back)

    flag = ", ".join(sorted(tired)) if tired else "No relievers flagged for heavy recent use"

    return f"""### {team_name} Bullpen — Last {days_back} Days

{md_table(headers, rows)}

**Caution arms (pitched yesterday, or 3+ appearances in window):** {flag}
"""

def pitcher_rows(sc, name, pitcher_id=None):
    if sc.empty or name == "TBD":
        return pd.DataFrame()

    if pitcher_id and "pitcher" in sc.columns:
        by_id = sc[sc["pitcher"] == pitcher_id]
        if not by_id.empty:
            return by_id

    if "player_name" not in sc.columns:
        return pd.DataFrame()

    parts = name.split()
    if len(parts) < 2:
        return pd.DataFrame()

    reverse = f"{parts[-1]}, {' '.join(parts[:-1])}"
    return sc[sc["player_name"].astype(str).str.lower() == reverse.lower()]

def event_stats(df):
    if df.empty or "events" not in df.columns:
        return 0, 0, 0, None, None, 0, 0, 0, 0

    events = df.dropna(subset=["events"])
    ab_events = events[~events["events"].isin(["walk", "intent_walk", "hit_by_pitch", "sac_bunt", "sac_fly"])]
    ab = len(ab_events)

    singles = (events["events"] == "single").sum()
    doubles = (events["events"] == "double").sum()
    triples = (events["events"] == "triple").sum()
    hrs = (events["events"] == "home_run").sum()

    hits = singles + doubles + triples + hrs
    tb = singles + 2 * doubles + 3 * triples + 4 * hrs

    avg = hits / ab if ab else 0
    slg = tb / ab if ab else 0
    iso = slg - avg if ab else 0

    woba = df["woba_value"].mean() if "woba_value" in df.columns else None
    xwoba = df["estimated_woba_using_speedangle"].mean() if "estimated_woba_using_speedangle" in df.columns else None
    xba = df["estimated_ba_using_speedangle"].mean() if "estimated_ba_using_speedangle" in df.columns else None
    xslg = df["estimated_slg_using_speedangle"].mean() if "estimated_slg_using_speedangle" in df.columns else None

    bbe = df["launch_speed"].dropna() if "launch_speed" in df.columns else pd.Series(dtype=float)
    hard = (bbe >= 95).mean() if len(bbe) else 0

    barrel = 0
    sweet = 0
    if "launch_speed_angle" in df.columns:
        lsa = df["launch_speed_angle"].dropna()
        barrel = (lsa == 6).mean() if len(lsa) else 0
    if "launch_angle" in df.columns:
        la = df["launch_angle"].dropna()
        sweet = ((la >= 8) & (la <= 32)).mean() if len(la) else 0

    swings = df[df["description"].isin(["swinging_strike", "swinging_strike_blocked", "foul", "foul_tip", "hit_into_play"])]
    whiffs = df[df["description"].isin(["swinging_strike", "swinging_strike_blocked"])]
    whiff = len(whiffs) / len(swings) if len(swings) else 0

    return avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, len(events), singles, hits, tb

def pitcher_profile(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)

    if p.empty:
        return md_table(["Stat", "Value"], [])

    events = p.dropna(subset=["events"])
    h = events["events"].isin(["single", "double", "triple", "home_run"]).sum()
    bb = events["events"].isin(["walk", "intent_walk"]).sum()
    k = events["events"].isin(["strikeout", "strikeout_double_play"]).sum()
    hr = (events["events"] == "home_run").sum()

    total_events = max(1, len(events))

    rows = [
        ["Sample Pitches", len(p)],
        ["Batted/Result Events", len(events)],
        ["Hits Allowed", h],
        ["Walks", bb],
        ["Strikeouts", k],
        ["Home Runs", hr],
        ["K Event Rate", pct(k / total_events)],
        ["BB Event Rate", pct(bb / total_events)],
        ["HR Event Rate", pct(hr / total_events)],
    ]

    return md_table(["Stat", "Value"], rows)

def arsenal(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    headers = ["Pitch", "Side", "Usage", "Pitches", "Hits", "1B", "TB", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]

    if p.empty or "pitch_type" not in p.columns or "stand" not in p.columns:
        return md_table(headers, [])

    rows = []
    total = max(1, len(p))

    for (pitch, side), g in p.groupby(["pitch_type", "stand"]):
        avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, pa, singles, hits, tb = event_stats(g)
        rows.append([
            pitch,
            f"vs {side}",
            pct(len(g) / total),
            len(g),
            fmt(avg),
            fmt(slg),
            fmt(iso),
            fmt(woba),
            fmt(xwoba),
            pct(barrel),
            pct(hard),
            pct(whiff),
        ])

    return md_table(headers, sorted(rows, key=lambda x: (x[0], x[1])))

def major_pitches(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    if p.empty or "pitch_type" not in p.columns:
        return []

    mix = p["pitch_type"].value_counts(normalize=True)
    return list(mix[mix >= 0.08].index)

def last_5_starts(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    headers = ["Date", "Pitches", "Hits", "BB", "K", "HR"]

    if p.empty or "game_date" not in p.columns:
        return md_table(headers, [])

    rows = []

    for game_date, g in p.groupby("game_date"):
        events = g.dropna(subset=["events"])
        hits = events["events"].isin(["single", "double", "triple", "home_run"]).sum()
        bb = events["events"].isin(["walk", "intent_walk"]).sum()
        k = events["events"].isin(["strikeout", "strikeout_double_play"]).sum()
        hr = (events["events"] == "home_run").sum()
        rows.append([game_date, len(g), hits, bb, k, hr])

    rows = sorted(rows, key=lambda x: str(x[0]), reverse=True)[:5]
    return md_table(headers, rows)

def team_hitter_pool(sc, team_code):
    if sc.empty or not team_code or "batter_team" not in sc.columns or "batter_name" not in sc.columns:
        return pd.DataFrame()

    t = sc[sc["batter_team"] == team_code]
    rows = []

    for name, g in t.groupby("batter_name"):
        if not name or pd.isna(name):
            continue

        avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, pa, singles, hits, tb = event_stats(g)

        if pa < 10:
            continue

        rows.append({
            "Hitter": str(name).title(),
            "PA": pa,
            "AVG": avg,
            "SLG": slg,
            "ISO": iso,
            "wOBA": woba if woba is not None else 0,
            "xwOBA": xwoba if xwoba is not None else 0,
            "Barrel%": barrel,
            "HardHit%": hard,
            "Whiff%": whiff,
        })

    df = pd.DataFrame(rows)

    if df.empty or "PA" not in df.columns:
        return pd.DataFrame()

    return df.sort_values(["PA", "wOBA"], ascending=[False, False]).head(12)

def hitter_pool_md(df):
    headers = ["Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]

    if df.empty:
        return md_table(headers, [])

    rows = []
    for _, r in df.iterrows():
        rows.append([
            r["Hitter"],
            int(r["PA"]),
            fmt(r["AVG"]),
            fmt(r["SLG"]),
            fmt(r["ISO"]),
            fmt(r["wOBA"]),
            fmt(r["xwOBA"]),
            pct(r["Barrel%"]),
            pct(r["HardHit%"]),
            pct(r["Whiff%"]),
        ])

    return md_table(headers, rows)

def hitter_vs_pitch_rows(sc, hitters, pitches):
    if sc.empty or hitters.empty or not pitches or "batter_name" not in sc.columns or "pitch_type" not in sc.columns:
        return []

    names = set(hitters["Hitter"].astype(str).str.lower())
    rows = []

    for pitch in pitches:
        pitch_df = sc[sc["pitch_type"] == pitch]

        for name in names:
            h = pitch_df[pitch_df["batter_name"].astype(str).str.lower() == name]

            if h.empty:
                continue

            avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, pa, singles, hits, tb = event_stats(h)

            if pa < 3:
                continue

            rows.append({
                "Pitch": pitch, "Hitter": name.title(), "PA": pa,
                "AVG": avg, "SLG": slg, "ISO": iso,
                "wOBA": woba if pd.notna(woba) else 0,
                "xwOBA": xwoba if pd.notna(xwoba) else 0,
            "xBA": xba if pd.notna(xba) else 0,
            "xSLG": xslg if pd.notna(xslg) else 0,
            "SweetSpot%": sweet,
                "Barrel%": barrel, "HardHit%": hard, "Whiff%": whiff,
            })

    rows.sort(key=lambda r: r["wOBA"], reverse=True)
    return rows[:30]

def hitter_vs_pitch(sc, hitters, pitches):
    headers = ["Pitch", "Hitter", "PA", "Hits", "1B", "TB", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]

    if sc.empty or hitters.empty or not pitches or "batter_name" not in sc.columns or "pitch_type" not in sc.columns:
        return md_table(headers, [])

    names = set(hitters["Hitter"].astype(str).str.lower())
    rows = []

    for pitch in pitches:
        pitch_df = sc[sc["pitch_type"] == pitch]

        for name in names:
            h = pitch_df[pitch_df["batter_name"].astype(str).str.lower() == name]

            if h.empty:
                continue

            avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, pa, singles, hits, tb = event_stats(h)

            if pa < 3:
                continue

            rows.append([
                pitch,
                name.title(),
                pa,
                fmt(avg),
                fmt(slg),
                fmt(iso),
                fmt(woba),
                fmt(xwoba),
                pct(barrel),
                pct(hard),
                pct(whiff),
            ])

    if not rows:
        return md_table(headers, [])

    rows = sorted(rows, key=lambda r: float(r[5]) if r[5] != "—" else 0, reverse=True)[:30]
    return md_table(headers, rows)

def game_dissection(sc, g, away_hitters, home_hitters):
    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    return f"""
## Projected Lineups

### Away Projected Lineup

{projected_lineup_table(away_hitters)}

### Home Projected Lineup

{projected_lineup_table(home_hitters)}

## Bullpen / Staff Context

### Away Staff

{bullpen_context_table(sc, g.get("away_code"))}

### Home Staff

{bullpen_context_table(sc, g.get("home_code"))}

## Final Game Dissection

- Away pitcher pitch mix to inspect: {", ".join(away_mix) if away_mix else "No current Statcast sample"}
- Home pitcher pitch mix to inspect: {", ".join(home_mix) if home_mix else "No current Statcast sample"}
- Home hitters should be checked against away SP pitch mix above.
- Away hitters should be checked against home SP pitch mix above.
- Lineup advantage: compare wOBA, xwOBA, ISO, Barrel%, HardHit%, Whiff%.
- Handedness advantage: use pitcher arsenal vs L/R tables.
- Bullpen fatigue: see Bullpen Fatigue Report above.
- Final MLB-LAB read: decide from pitch mix, hitter pool, L/R damage, current form, and bullpen fatigue.
"""

def game_card(i, g, sc):
    away_hitters = team_hitter_pool(sc, g["away_code"])
    home_hitters = team_hitter_pool(sc, g["home_code"])

    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    return f"""
---

# {i}. {g['game']}

## Game Context

{md_table(["Field", "Value"], [
    ["Park", g["park"]],
    ["Time", g["time"]],
    ["Away Team", g["away_team"]],
    ["Home Team", g["home_team"]],
    ["Away Probable Pitcher", g["away_sp"]],
    ["Home Probable Pitcher", g["home_sp"]],
])}

## Away Starting Pitcher: {g['away_sp']}

### Pitcher Profile

{pitcher_profile(sc, g["away_sp"], g.get("away_sp_id"))}

#### Last 5 Starts

{last_five_starts_table(sc, g.get("away_sp_id"))}

### Full Pitch Arsenal vs L/R

{arsenal(sc, g["away_sp"], g.get("away_sp_id"))}

### Last 5 Starts / Appearances

{last_5_starts(sc, g["away_sp"], g.get("away_sp_id"))}

## Home Hitters vs Away SP Pitch Mix

{hitter_vs_pitch(sc, home_hitters, away_mix)}

## Home Starting Pitcher: {g['home_sp']}

### Pitcher Profile

{pitcher_profile(sc, g["home_sp"], g.get("home_sp_id"))}

#### Last 5 Starts

{last_five_starts_table(sc, g.get("home_sp_id"))}

### Full Pitch Arsenal vs L/R

{arsenal(sc, g["home_sp"], g.get("home_sp_id"))}

### Last 5 Starts / Appearances

{last_5_starts(sc, g["home_sp"], g.get("home_sp_id"))}

## Away Hitters vs Home SP Pitch Mix

{hitter_vs_pitch(sc, away_hitters, home_mix)}

## {g['away_team']} Team Hitter Pool

{hitter_pool_md(away_hitters)}

## {g['home_team']} Team Hitter Pool

{hitter_pool_md(home_hitters)}

## Bullpen Fatigue Report

{bullpen_context(g.get("away_team_id"), g["away_team"])}

{bullpen_context(g.get("home_team_id"), g["home_team"])}

{game_dissection(sc, g, away_hitters, home_hitters)}
"""

PCT_COLS = {"Barrel%", "HardHit%", "Whiff%", "K Event Rate", "BB Event Rate", "HR Event Rate", "Usage"}
DEC_COLS = {"AVG", "SLG", "ISO", "wOBA", "xwOBA"}

def apply_color_scale(ws, col_letter, first_row, last_row, invert=False):
    if last_row < first_row:
        return
    lo, hi = ("FFC7CE", "C6EFCE") if not invert else ("C6EFCE", "FFC7CE")
    rule = ColorScaleRule(
        start_type="min", start_color=lo,
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color=hi,
    )
    ws.conditional_formatting.add(f"{col_letter}{first_row}:{col_letter}{last_row}", rule)

def style_table(ws, start_row, start_col, headers, rows, pct_cols, dec_cols, color_cols=None):
    """Writes a header+rows table at (start_row, start_col) with banding,
    number formats, and per-column color scales. Returns the row after the table."""
    color_cols = color_cols if color_cols is not None else (PCT_COLS | DEC_COLS | {"K%", "BB%", "HR Event Rate"})

    for c, h in enumerate(headers):
        col = start_col + c
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows):
        excel_row = start_row + 1 + r
        for c, h in enumerate(headers):
            col = start_col + c
            val = row.get(h, "")
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            if h in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
            elif h in dec_cols and isinstance(val, (int, float)):
                cell.number_format = "0.000"

    last_row = start_row + len(rows)
    if rows:
        for c, h in enumerate(headers):
            if h in color_cols:
                apply_color_scale(ws, start_col + c, start_row + 1, last_row, invert=(h in GOOD_LOW_STATS))

    widths = [7, 38, 22, 20, 7, 18, 6, 7, 7, 7, 7, 7, 8, 9, 8]
    for c, w in enumerate(widths[:len(headers)], start_col):
        ws.column_dimensions[get_column_letter(c)].width = w

    return last_row + 1


def section_header(ws, row, col, text, span):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return row + 1

def build_slate_sheet(wb, games):
    ws = wb.create_sheet("Slate")
    headers = ["#", "Game", "Park", "Away SP", "Home SP", "Time (UTC)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, g in enumerate(games, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=g["game"])
        ws.cell(row=i + 1, column=3, value=g["park"])
        ws.cell(row=i + 1, column=4, value=g["away_sp"])
        ws.cell(row=i + 1, column=5, value=g["home_sp"])
        ws.cell(row=i + 1, column=6, value=g["time"])
        for c in range(1, 7):
            ws.cell(row=i + 1, column=c).font = BODY_FONT

    widths = [4, 40, 24, 22, 22, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(games) + 1}"
    return ws

def build_game_sheet(wb, i, g, sc):
    away_hitters = team_hitter_pool(sc, g["away_code"])
    home_hitters = team_hitter_pool(sc, g["home_code"])
    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    home_vs_away = hitter_vs_pitch_rows(sc, home_hitters, away_mix)
    away_vs_home = hitter_vs_pitch_rows(sc, away_hitters, home_mix)

    away_arsenal = arsenal_rows(sc, g["away_sp"], g.get("away_sp_id"))
    home_arsenal = arsenal_rows(sc, g["home_sp"], g.get("home_sp_id"))

    away_bullpen, away_tired = bullpen_usage(g.get("away_team_id"), 4)
    home_bullpen, home_tired = bullpen_usage(g.get("home_team_id"), 4)

    tab = f"G{i} {g['away_code']}@{g['home_code']}"
    ws = wb.create_sheet(tab[:31])

    ws["A1"] = g["game"]
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{g['park']}  |  {g['time']}  |  Away SP: {g['away_sp']}  |  Home SP: {g['home_sp']}"
    ws["A2"].font = SUBHEAD_FONT

    arsenal_headers = ["Pitch", "Side", "Usage", "Pitches", "Hits", "1B", "TB", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]
    matchup_headers = ["Pitch", "Hitter", "PA", "Hits", "1B", "TB", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]
    pool_headers = ["Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]
    bullpen_headers = ["Reliever", "Date", "IP", "Pitches"]

    row = 4
    # --- AWAY SP arsenal, then HOME hitters vs that arsenal (the actual matchup) ---
    row = section_header(ws, row, 1, f"{g['away_sp']} ({g['away_team']}) — Pitch Arsenal", len(arsenal_headers))
    row = style_table(ws, row, 1, arsenal_headers, away_arsenal, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['home_team']} Hitters vs {g['away_sp']} Pitch Mix", len(matchup_headers))
    row = style_table(ws, row, 1, matchup_headers, home_vs_away, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['home_team']} Hitter Pool (Season)", len(pool_headers))
    row = style_table(ws, row, 1, pool_headers, home_hitters.to_dict("records"), PCT_COLS, DEC_COLS) + 2

    # --- HOME SP arsenal, then AWAY hitters vs that arsenal ---
    row = section_header(ws, row, 1, f"{g['home_sp']} ({g['home_team']}) — Pitch Arsenal", len(arsenal_headers))
    row = style_table(ws, row, 1, arsenal_headers, home_arsenal, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['away_team']} Hitters vs {g['home_sp']} Pitch Mix", len(matchup_headers))
    row = style_table(ws, row, 1, matchup_headers, away_vs_home, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['away_team']} Hitter Pool (Season)", len(pool_headers))
    row = style_table(ws, row, 1, pool_headers, away_hitters.to_dict("records"), PCT_COLS, DEC_COLS) + 2

    # --- Bullpen fatigue, both teams ---
    row = section_header(ws, row, 1, f"{g['away_team']} Bullpen — Last 4 Days", len(bullpen_headers))
    away_bp_rows = [{"Reliever": r[0], "Date": r[1], "IP": r[2], "Pitches": r[3]} for r in away_bullpen]
    row = style_table(ws, row, 1, bullpen_headers, away_bp_rows, set(), set(), color_cols=set())
    if away_tired:
        ws.cell(row=row, column=1, value=f"Caution arms: {', '.join(sorted(away_tired))}").font = Font(italic=True, color="C00000")
    row += 2

    row = section_header(ws, row, 1, f"{g['home_team']} Bullpen — Last 4 Days", len(bullpen_headers))
    home_bp_rows = [{"Reliever": r[0], "Date": r[1], "IP": r[2], "Pitches": r[3]} for r in home_bullpen]
    row = style_table(ws, row, 1, bullpen_headers, home_bp_rows, set(), set(), color_cols=set())
    if home_tired:
        ws.cell(row=row, column=1, value=f"Caution arms: {', '.join(sorted(home_tired))}").font = Font(italic=True, color="C00000")

    widths = [18, 22, 8, 8, 8, 8, 8, 8, 9, 10, 9, 9]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws

def arsenal_rows(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    if p.empty or "pitch_type" not in p.columns or "stand" not in p.columns:
        return []

    rows = []
    total = max(1, len(p))
    for (pitch, side), g in p.groupby(["pitch_type", "stand"]):
        avg, slg, iso, woba, xwoba, xba, xslg, barrel, hard, sweet, whiff, pa, singles, hits, tb = event_stats(g)
        rows.append({
            "Pitch": pitch, "Side": f"vs {side}", "Usage": len(g) / total, "Pitches": len(g),
            "Hits": hits,
            "1B": singles,
            "TB": tb,
            "AVG": avg, "SLG": slg, "ISO": iso,
            "wOBA": woba if pd.notna(woba) else 0,
            "xwOBA": xwoba if pd.notna(xwoba) else 0,
            "xBA": xba if pd.notna(xba) else 0,
            "xSLG": xslg if pd.notna(xslg) else 0,
            "SweetSpot%": sweet,
            "Barrel%": barrel, "HardHit%": hard, "Whiff%": whiff,
        })
    return sorted(rows, key=lambda r: (r["Pitch"], r["Side"]))


def build_top_plays_sheet(wb, games, sc):
    ws = wb.create_sheet("Top Plays", 1)
    headers = ["Category", "Rank", "Score", "Tier", "Game", "Team", "Hitter", "Opp SP", "Pitch", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%", "Hits", "1B", "TB"]

    def val(r, k):
        try:
            return float(r.get(k, 0) or 0)
        except Exception:
            return 0

    def score(r, cat):
        avg, slg, iso = val(r,"AVG"), val(r,"SLG"), val(r,"ISO")
        woba, xwoba, xba, xslg = val(r,"wOBA"), val(r,"xwOBA"), val(r,"xBA"), val(r,"xSLG")
        sweet, barrel, hard, whiff = val(r,"SweetSpot%"), val(r,"Barrel%"), val(r,"HardHit%"), val(r,"Whiff%")
        hits, singles, tb = val(r,"Hits"), val(r,"1B"), val(r,"TB")

        if cat == "Hits":
            return xba*30 + avg*25 + xwoba*20 + woba*15 + (1-whiff)*10
        if cat == "Singles":
            return avg*30 + xba*25 + singles*2 + (1-whiff)*15 + sweet*10
        if cat == "Total Bases":
            return xslg*30 + slg*20 + iso*20 + tb*1.5 + barrel*15 + hard*10
        if cat == "HRR":
            return woba*25 + xwoba*25 + xba*15 + hits*1.2 + tb*1 + hard*10
        if cat == "Home Runs":
            return iso*30 + xslg*25 + barrel*25 + hard*15 + slg*10
        return 0

    rows = []
    for i, g in enumerate(games, 1):
        away_hitters = team_hitter_pool(sc, g["away_code"])
        home_hitters = team_hitter_pool(sc, g["home_code"])
        away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
        home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

        for r in hitter_vs_pitch_rows(sc, home_hitters, away_mix):
            r.update({"Game": g["game"], "Team": g["home_team"], "Opp SP": g["away_sp"]})
            rows.append(r)
        for r in hitter_vs_pitch_rows(sc, away_hitters, home_mix):
            r.update({"Game": g["game"], "Team": g["away_team"], "Opp SP": g["home_sp"]})
            rows.append(r)

    rownum = 1
    categories = ["Hits", "Singles", "Total Bases", "HRR", "Home Runs"]

    for cat in categories:
        ws.cell(row=rownum, column=1, value=f"TOP 5 {cat.upper()}")
        ws.cell(row=rownum, column=1).font = HEADER_FONT
        ws.cell(row=rownum, column=1).fill = HEADER_FILL
        rownum += 1

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=rownum, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        rownum += 1

        ranked = []
        for r in rows:
            scv = score(r, cat)
            tier = "T1" if scv >= 35 else "T2" if scv >= 25 else "T3"
            ranked.append((scv, tier, r))

        for rank, (scv, tier, r) in enumerate(sorted(ranked, key=lambda x: x[0], reverse=True)[:5], 1):
            vals = [cat, rank, round(scv, 3), tier, r.get("Game"), r.get("Team"), r.get("Hitter"), r.get("Opp SP"), r.get("Pitch"),
                    r.get("PA"), r.get("AVG"), r.get("SLG"), r.get("ISO"), r.get("wOBA"), r.get("xwOBA"), r.get("xBA"), r.get("xSLG"),
                    r.get("SweetSpot%"), r.get("Barrel%"), r.get("HardHit%"), r.get("Whiff%"), r.get("Hits"), r.get("1B"), r.get("TB")]
            for c, v in enumerate(vals, 1):
                ws.cell(row=rownum, column=c, value=v)
            rownum += 1
        rownum += 2

    ws.freeze_panes = "A3"
    for c in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 15


def build_category_boards_sheet(wb, games, sc):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=9)
    BODY_FONT    = Font(size=9)
    SECTION_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    T1_FILL      = PatternFill("solid", fgColor="C6EFCE")
    T2_FILL      = PatternFill("solid", fgColor="FFEB9C")
    T3_FILL      = PatternFill("solid", fgColor="FBBF72")

    COLS = ["Category","Rank","Hitter","Team","Game","Opp SP","Pitch","PA","Hits","1B","TB","AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","SweetSpot%","Barrel%","HardHit%","Whiff%"]

    CATEGORIES = [
        ("Hits Leaders", ["Hits", "xBA", "AVG"]),
        ("Singles Leaders", ["1B", "AVG", "xBA"]),
        ("Total Bases Leaders", ["TB", "xSLG", "SLG", "ISO"]),
        ("HRR Leaders", ["wOBA", "xwOBA", "Hits", "TB", "HardHit%"]),
        ("Home Run Leaders", ["ISO", "xSLG", "Barrel%", "HardHit%"]),
    ]

    all_rows = []
    for g in games:
        away_hitters = team_hitter_pool(sc, g["away_code"])
        home_hitters = team_hitter_pool(sc, g["home_code"])
        away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
        home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

        for r in hitter_vs_pitch_rows(sc, home_hitters, away_mix):
            all_rows.append({**r, "Game": g.get("game", ""), "Team": g.get("home_team", ""), "Opp SP": g.get("away_sp", "")})
        for r in hitter_vs_pitch_rows(sc, away_hitters, home_mix):
            all_rows.append({**r, "Game": g.get("game", ""), "Team": g.get("away_team", ""), "Opp SP": g.get("home_sp", "")})

    ws = wb.create_sheet("Category Boards", 1)
    rownum = 1

    for section_label, sort_keys in CATEGORIES:
        primary_key = sort_keys[0]

        def composite(r, keys=sort_keys):
            vals = []
            for k in keys:
                try:
                    vals.append(float(r.get(k, 0) or 0))
                except Exception:
                    pass
            return sum(vals) / len(vals) if vals else 0

        top20 = sorted(all_rows, key=composite, reverse=True)[:20]

        ws.cell(row=rownum, column=1, value=section_label).font = SECTION_FONT
        ws.cell(row=rownum, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=rownum, start_column=1, end_row=rownum, end_column=len(COLS))
        rownum += 1

        for c, h in enumerate(COLS, 1):
            cell = ws.cell(row=rownum, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        rownum += 1

        for rank, r in enumerate(top20, 1):
            try:
                scv = float(r.get(primary_key, 0) or 0)
            except Exception:
                scv = 0
            fill = T1_FILL if scv >= 10 else T2_FILL if scv >= 5 else T3_FILL if primary_key in ("Hits","1B","TB") else T1_FILL if scv >= 0.370 else T2_FILL if scv >= 0.320 else T3_FILL
            row_vals = {"Category": section_label, "Rank": rank, **r}

            for c, h in enumerate(COLS, 1):
                val = row_vals.get(h, "")
                cell = ws.cell(row=rownum, column=c, value=val)
                cell.font = BODY_FONT
                cell.fill = fill
                if isinstance(val, float):
                    
try:
    cell.value = float(cell.value)
except:
    pass
cell.number_format = "0.000"

            rownum += 1

        rownum += 2

    ws.freeze_panes = "H3"
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max_len + 2, 24)



def build_betting_command_center(wb, games, sc):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from pathlib import Path
    import html

    # remove old workflow tabs
    for name in ["🚨 Command Center", "Category Explorer", "Top Plays", "Category Boards"]:
        if name in wb.sheetnames:
            del wb[name]

    DARK = "0D1117"
    ROW1 = "161B22"
    ROW2 = "1C2128"
    BLUE = "1F4E79"
    GREEN = "1A9C3E"
    YELLOW = "F0B429"
    ORANGE = "E07B39"
    RED = "C0392B"
    WHITE = "FFFFFF"
    GOLD = "D4A017"

    thin = Side(style="thin", color="30363D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(sz=9, bold=False, color=WHITE): return Font(name="Calibri", size=sz, bold=bold, color=color)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def num(r, k):
        try:
            return float(r.get(k, 0) or 0)
        except Exception:
            return 0.0

    def rate(n, d):
        return n / d if d else 0

    def pct_rank(values, value):
        values = sorted([v for v in values if v is not None])
        if not values:
            return 0
        below = sum(1 for v in values if v <= value)
        return below / len(values)

    rows = []
    for g in games:
        away_h = team_hitter_pool(sc, g["away_code"])
        home_h = team_hitter_pool(sc, g["home_code"])
        away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
        home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

        for r in hitter_vs_pitch_rows(sc, home_h, away_mix):
            rows.append({**r, "Game": g.get("game", ""), "Team": g.get("home_team", ""), "Opp SP": g.get("away_sp", "")})
        for r in hitter_vs_pitch_rows(sc, away_h, home_mix):
            rows.append({**r, "Game": g.get("game", ""), "Team": g.get("away_team", ""), "Opp SP": g.get("home_sp", "")})

    # percentile pools
    metrics = ["PA", "Hits", "1B", "TB", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]
    pools = {m: [num(r, m) for r in rows] for m in metrics}

    def pr(r, k, invert=False):
        v = num(r, k)
        p = pct_rank(pools.get(k, []), v)
        return 1 - p if invert else p

    def score(r, cat):
        if cat == "Hits":
            return 100 * (pr(r,"xBA")*.30 + pr(r,"AVG")*.25 + pr(r,"xwOBA")*.20 + pr(r,"wOBA")*.15 + pr(r,"Whiff%", True)*.10)
        if cat == "Singles":
            return 100 * (pr(r,"AVG")*.30 + pr(r,"xBA")*.25 + pr(r,"1B")*.20 + pr(r,"Whiff%", True)*.15 + pr(r,"SweetSpot%")*.10)
        if cat == "Total Bases":
            return 100 * (pr(r,"xSLG")*.30 + pr(r,"SLG")*.20 + pr(r,"ISO")*.20 + pr(r,"TB")*.15 + pr(r,"Barrel%")*.10 + pr(r,"HardHit%")*.05)
        if cat == "HRR":
            return 100 * (pr(r,"wOBA")*.20 + pr(r,"xwOBA")*.20 + pr(r,"xBA")*.15 + pr(r,"Hits")*.15 + pr(r,"TB")*.15 + pr(r,"HardHit%")*.15)
        if cat == "Home Runs":
            return 100 * (pr(r,"ISO")*.25 + pr(r,"xSLG")*.25 + pr(r,"Barrel%")*.25 + pr(r,"HardHit%")*.15 + pr(r,"SLG")*.10)
        return 0

    def tier(v):
        return "A" if v >= 80 else "B" if v >= 65 else "C" if v >= 50 else "D"

    def tfill(t):
        return fill(GREEN if t == "A" else YELLOW if t == "B" else ORANGE if t == "C" else RED)

    def risk(r):
        risks = []
        if num(r,"PA") < 15: risks.append("Low PA")
        if num(r,"Whiff%") > 0.30: risks.append("High Whiff")
        if num(r,"xBA") < 0.230: risks.append("Weak xBA")
        if num(r,"Barrel%") < 0.04 and num(r,"ISO") < 0.140: risks.append("Low Power")
        return "None" if not risks else " + ".join(risks[:2])

    def why(r, cat):
        if cat == "Hits": return f"xBA {num(r,'xBA'):.3f}, AVG {num(r,'AVG'):.3f}, Whiff {num(r,'Whiff%'):.1%}"
        if cat == "Singles": return f"1B {int(num(r,'1B'))}, xBA {num(r,'xBA'):.3f}, Sweet {num(r,'SweetSpot%'):.1%}"
        if cat == "Total Bases": return f"xSLG {num(r,'xSLG'):.3f}, TB {int(num(r,'TB'))}, ISO {num(r,'ISO'):.3f}"
        if cat == "HRR": return f"wOBA {num(r,'wOBA'):.3f}, TB {int(num(r,'TB'))}, HardHit {num(r,'HardHit%'):.1%}"
        return f"ISO {num(r,'ISO'):.3f}, Barrel {num(r,'Barrel%'):.1%}, xSLG {num(r,'xSLG'):.3f}"

    def best_unique(cat, n):
        seen = set()
        out = []
        for r in sorted(rows, key=lambda x: score(x, cat), reverse=True):
            key = (r.get("Hitter"), r.get("Team"), cat)
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
            if len(out) >= n:
                break
        return out

    cats = [("🔥 Hits", "Hits"), ("🎯 Singles", "Singles"), ("💥 Total Bases", "Total Bases"), ("⚡ HRR", "HRR"), ("💣 Home Runs", "Home Runs")]

    # Command Center: clean decision page
    ws = wb.create_sheet("🚨 Command Center", 1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "⚾ MLB-LAB COMMAND CENTER"
    ws["A1"].font = font(16, True, GOLD)
    ws["A1"].fill = fill(DARK)
    ws.merge_cells("A1:J1")

    headers = ["Rank", "Player", "Team", "Game", "Opp SP", "Best Market", "Score", "Tier", "Why", "Risk"]
    rnum = 3
    pdf_rows = []

    for label, cat in cats:
        ws.cell(rnum, 1, label).fill = fill(BLUE)
        ws.cell(rnum, 1).font = font(11, True)
        ws.merge_cells(start_row=rnum, start_column=1, end_row=rnum, end_column=10)
        rnum += 1

        for c, h in enumerate(headers, 1):
            cell = ws.cell(rnum, c, h)
            cell.fill = fill(DARK)
            cell.font = font(9, True)
            cell.alignment = center
            cell.border = border
        rnum += 1

        for rank, row in enumerate(best_unique(cat, 5), 1):
            scv = score(row, cat)
            tr = tier(scv)
            vals = [rank, row.get("Hitter",""), row.get("Team",""), row.get("Game",""), row.get("Opp SP",""), label, round(scv,1), tr, why(row,cat), risk(row)]
            pdf_rows.append(vals)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rnum, c, v)
                cell.border = border
                cell.font = font(9, c in [1,2,6,7,8])
                cell.alignment = left if c in [2,4,9,10] else center
                cell.fill = tfill(tr) if c == 8 else fill(ROW1 if rank % 2 else ROW2)
            rnum += 1
        rnum += 2

    for c, w in enumerate([7,22,18,34,18,16,10,8,46,24], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 90

    # Category Explorer: research page
    bs = wb.create_sheet("Category Explorer", 2)
    bs.sheet_view.showGridLines = False
    bs.sheet_properties.tabColor = BLUE
    cols = ["Category","Rank","Hitter","Team","Game","Opp SP","Pitch","Score","Tier","Risk","PA","Hits","1B","TB","AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","SweetSpot%","Barrel%","HardHit%","Whiff%"]
    rnum = 1

    for label, cat in cats:
        bs.cell(rnum, 1, f"{label.upper()} BOARD — TOP 20").fill = fill(BLUE)
        bs.cell(rnum, 1).font = font(11, True)
        bs.merge_cells(start_row=rnum, start_column=1, end_row=rnum, end_column=len(cols))
        rnum += 1

        for c, h in enumerate(cols, 1):
            cell = bs.cell(rnum, c, h)
            cell.fill = fill(DARK)
            cell.font = font(8, True)
            cell.alignment = center
            cell.border = border
        rnum += 1

        for rank, row in enumerate(best_unique(cat, 20), 1):
            scv = score(row, cat)
            tr = tier(scv)
            vals = [label, rank, row.get("Hitter",""), row.get("Team",""), row.get("Game",""), row.get("Opp SP",""), row.get("Pitch",""), round(scv,1), tr, risk(row),
                    row.get("PA",""), row.get("Hits",""), row.get("1B",""), row.get("TB",""), row.get("AVG",""), row.get("SLG",""), row.get("ISO",""),
                    row.get("wOBA",""), row.get("xwOBA",""), row.get("xBA",""), row.get("xSLG",""), row.get("SweetSpot%",""), row.get("Barrel%",""),
                    row.get("HardHit%",""), row.get("Whiff%","")]
            for c, v in enumerate(vals, 1):
                cell = bs.cell(rnum, c, v)
                cell.border = border
                cell.font = font(8, c in [2,3,8,9])
                cell.alignment = center if c not in [3,4,5,6,10] else left
                cell.fill = tfill(tr) if c == 9 else fill(ROW1 if rank % 2 else ROW2)
                if isinstance(v, float):
                    if c >= 15:
                        
try:
    cell.value = float(cell.value)
except:
    pass
cell.number_format = "0.000"

                    else:
                        cell.number_format = "0.0"
            rnum += 1
        rnum += 2

    widths = [16,6,20,16,30,18,8,8,8,22,8,8,8,8,8,8,8,8,8,8,8,11,9,10,9]
    for c, w in enumerate(widths, 1):
        bs.column_dimensions[get_column_letter(c)].width = w
    bs.freeze_panes = "K3"
    bs.sheet_view.zoomScale = 85

    # polish existing game tabs: fix cramped metric columns
    for sheet in wb.worksheets:
        if sheet.title in ["🚨 Command Center", "Category Explorer"]:
            continue
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 90
        for col_idx in range(1, min(sheet.max_column, 28) + 1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = 12
        for col_idx in range(1, min(sheet.max_column, 8) + 1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = max(sheet.column_dimensions[letter].width or 12, 16)
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 220), max_col=min(sheet.max_column, 28)):
            for cell in row:
                if isinstance(cell.value, float):
                    
try:
    cell.value = float(cell.value)
except:
    pass
cell.number_format = "0.000"
 if abs(cell.value) < 1 else "0.0"

    # one-page HTML report for printing/exporting
    report_path = Path("reports/mlb-lab-daily-command-center.html")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    html_rows = "".join(
        "<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in row) + "</tr>"
        for row in pdf_rows
    )
    report_path.write_text(f"""
<html><head><meta charset='utf-8'>
<style>
body {{ font-family: Arial; background:#0D1117; color:white; }}
h1 {{ color:#D4A017; }}
table {{ border-collapse:collapse; width:100%; font-size:12px; }}
th {{ background:#1F4E79; padding:6px; }}
td {{ border:1px solid #30363D; padding:5px; }}
tr:nth-child(even) {{ background:#161B22; }}
tr:nth-child(odd) {{ background:#1C2128; }}
</style></head>
<body>
<h1>⚾ MLB-LAB Daily Command Center</h1>
<table>
<tr>{''.join(f'<th>{h}</th>' for h in headers)}</tr>
{html_rows}
</table>
</body></html>
""", encoding="utf-8")
    print(f"Updated printable report: {report_path}")


def build_category_boards_sheet(wb, games, sc):
    build_betting_command_center(wb, games, sc)


def mlb_lab_apply_visual_cohesion(wb):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK="0D1117"; NAVY="1F4E79"; ROW1="F7F9FB"; ROW2="FFFFFF"
    GREEN="C6EFCE"; YELLOW="FFEB9C"; RED="F4CCCC"; ORANGE="FCE4D6"; WHITE="FFFFFF"
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c): return PatternFill("solid", fgColor=c)

    good_high = {"AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","SweetSpot%","Barrel%","HardHit%","Score"}
    good_low = {"Whiff%"}

    def color_for(header, val):
        try: v = float(val)
        except Exception: return None
        h = str(header)
        if h in good_low:
            return fill(GREEN if v <= .18 else YELLOW if v <= .28 else RED)
        if h in good_high:
            if h == "Score": return fill(GREEN if v >= 80 else YELLOW if v >= 65 else ORANGE if v >= 50 else RED)
            return fill(GREEN if v >= .370 else YELLOW if v >= .300 else RED)
        return None

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

        if ws.title == "Slate":
            ws.sheet_properties.tabColor = "1A9C3E"
            ws.freeze_panes = "A2"
        elif "Command Center" in ws.title:
            ws.sheet_properties.tabColor = "1A9C3E"
            ws.freeze_panes = "A3"
        elif ws.title == "Category Explorer":
            ws.sheet_properties.tabColor = "2E75B6"
            ws.freeze_panes = "K3"
        elif ws.title == "Master Matchups":
            ws.sheet_properties.tabColor = "D4A017"
            ws.freeze_panes = "A2"
        else:
            ws.sheet_properties.tabColor = "1F4E79"
            ws.freeze_panes = "A4"

        max_row = min(ws.max_row, 350)
        max_col = min(ws.max_column, 28)

        header_rows = set()
        for r in range(1, max_row + 1):
            vals = [ws.cell(r,c).value for c in range(1,max_col+1)]
            text = " ".join(str(v) for v in vals if v is not None)
            hits = sum(1 for x in vals if str(x) in ["Pitch","Hitter","PA","AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","Barrel%","HardHit%","Whiff%"])
            if hits >= 3:
                header_rows.add(r)
            if "Pitch Arsenal" in text or "Hitter Pool" in text or "Pitch Mix" in text or "Bullpen" in text:
                for c in range(1,max_col+1):
                    cell=ws.cell(r,c)
                    cell.fill=fill(NAVY)
                    cell.font=Font(color=WHITE,bold=True,size=9)
                    cell.alignment=Alignment(horizontal="center")
                    cell.border=border

        for r in range(1, max_row + 1):
            current_headers = {}
            last_header = max([hr for hr in header_rows if hr <= r], default=None)
            if last_header:
                current_headers = {c: ws.cell(last_header,c).value for c in range(1,max_col+1)}

            for c in range(1,max_col+1):
                cell = ws.cell(r,c)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if r in header_rows:
                    cell.fill = fill(DARK)
                    cell.font = Font(color=WHITE,bold=True,size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.value is not None:
                    cell.font = Font(size=8)
                    if r % 2 == 0 and ws.title not in ["🚨 Command Center","Category Explorer"]:
                        cell.fill = fill(ROW1)
                    h = current_headers.get(c)
                    cf = color_for(h, cell.value)
                    if cf:
                        cell.fill = cf
                    if isinstance(cell.value, float):
                        if str(h).endswith("%"):
                            cell.number_format = "0.0%"
                        else:
                            
try:
    cell.value = float(cell.value)
except:
    pass
cell.number_format = "0.000"
 if abs(cell.value) < 1 else "0.0"

        preferred = {
            "Pitch":8,"Side":8,"Usage":9,"Pitches":9,"Hitter":20,"Player":20,"Team":18,
            "Game":32,"Opp SP":18,"PA":8,"Hits":8,"1B":8,"TB":8,
            "AVG":8,"SLG":8,"ISO":8,"wOBA":8,"xwOBA":8,"xBA":8,"xSLG":8,
            "SweetSpot%":11,"Barrel%":9,"HardHit%":10,"Whiff%":9,
            "Why":44,"Risk":22,"Score":8,"Tier":7
        }

        for c in range(1,max_col+1):
            header = None
            for hr in sorted(header_rows):
                if ws.cell(hr,c).value:
                    header = str(ws.cell(hr,c).value)
                    break
            letter = get_column_letter(c)
            ws.column_dimensions[letter].width = preferred.get(header, 12)

        ws.row_dimensions[1].height = 22

    print("✅ visual cohesion applied: colors, widths, freeze panes, metric heatmaps")


def mlb_lab_final_product_polish(wb):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY="1F4E79"; BLUE="D9EAF7"; LIGHT="F7FBFF"; WHITE="FFFFFF"
    TEXT="111827"; GREEN="C6EFCE"; YELLOW="FFEB9C"; ORANGE="F4B183"; RED="F4CCCC"
    BORDER="D9E2EC"

    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    def fill(c): return PatternFill("solid", fgColor=c)
    def f(size=9,bold=False,color=TEXT): return Font(name="Calibri", size=size, bold=bold, color=color)

    def tier_fill(v):
        t = str(v or "").upper()
        return fill(GREEN if "A" in t else YELLOW if "B" in t else ORANGE if "C" in t else RED)

    def metric_fill(header, value):
        try: v = float(value)
        except Exception: return None
        h = str(header or "")
        if h == "Whiff%":
            return fill(GREEN if v <= .18 else YELLOW if v <= .28 else RED)
        if h in ["AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","SweetSpot%","Barrel%","HardHit%"]:
            return fill(GREEN if v >= .370 else YELLOW if v >= .300 else RED)
        if h == "Score":
            return fill(GREEN if v >= 80 else YELLOW if v >= 65 else ORANGE if v >= 50 else RED)
        return None

    # Collect Category Explorer rows for game summaries
    explorer_rows = []
    if "Category Explorer" in wb.sheetnames:
        ws = wb["Category Explorer"]
        headers = {}
        for r in range(1, ws.max_row + 1):
            vals = [ws.cell(r,c).value for c in range(1, min(ws.max_column,25)+1)]
            if "Category" in vals and "Hitter" in vals:
                headers = {ws.cell(r,c).value:c for c in range(1, min(ws.max_column,25)+1)}
                continue
            if headers and ws.cell(r, headers.get("Hitter", 0)).value:
                row = {k: ws.cell(r,c).value for k,c in headers.items()}
                explorer_rows.append(row)

    def top_for_game(game_name, market, n=3):
        out = []
        for r in explorer_rows:
            if market not in str(r.get("Category","")):
                continue
            if game_name and game_name not in str(r.get("Game","")):
                continue
            out.append(r)
        out = sorted(out, key=lambda x: float(x.get("Score") or 0), reverse=True)
        return out[:n]

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

        # LIGHT READABLE THEME — no black
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500), max_col=min(ws.max_column, 30)):
            for cell in row:
                cell.border = border
                cell.font = f(8)
                cell.alignment = Alignment(vertical="center")
                if cell.value is None:
                    continue

        # Header / section styling
        for r in range(1, min(ws.max_row,500)+1):
            vals = [ws.cell(r,c).value for c in range(1,min(ws.max_column,30)+1)]
            text = " ".join(str(v) for v in vals if v is not None)

            is_section = any(x in text for x in ["BOARD", "Pitch Arsenal", "Pitch Mix", "Hitter Pool", "Bullpen", "COMMAND CENTER"])
            header_hits = sum(1 for v in vals if str(v) in ["Rank","Player","Hitter","Team","Game","Opp SP","Pitch","Score","Tier","Risk","AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","Barrel%","HardHit%","Whiff%"])

            if is_section:
                for c in range(1,min(ws.max_column,30)+1):
                    cell=ws.cell(r,c)
                    cell.fill=fill(NAVY)
                    cell.font=f(9,True,WHITE)
                    cell.alignment=Alignment(horizontal="center",vertical="center")
            elif header_hits >= 3:
                for c in range(1,min(ws.max_column,30)+1):
                    cell=ws.cell(r,c)
                    cell.fill=fill(BLUE)
                    cell.font=f(8,True,TEXT)
                    cell.alignment=Alignment(horizontal="center",vertical="center")
            else:
                for c in range(1,min(ws.max_column,30)+1):
                    cell=ws.cell(r,c)
                    if cell.value is not None and ws.title not in ["🚨 Command Center","Category Explorer"]:
                        cell.fill=fill(LIGHT if r % 2 == 0 else WHITE)

        # Metric heatmaps by nearest header above
        header_row = {}
        for r in range(1, min(ws.max_row,500)+1):
            vals = [ws.cell(r,c).value for c in range(1,min(ws.max_column,30)+1)]
            if sum(1 for v in vals if str(v) in ["AVG","SLG","ISO","wOBA","xwOBA","xBA","xSLG","Barrel%","HardHit%","Whiff%","Score","Tier"]) >= 3:
                header_row = {c: ws.cell(r,c).value for c in range(1,min(ws.max_column,30)+1)}
                continue

            for c,h in header_row.items():
                cell = ws.cell(r,c)
                if str(h) == "Tier":
                    cell.fill = tier_fill(cell.value)
                    cell.font = f(8,True,TEXT)
                else:
                    mf = metric_fill(h, cell.value)
                    if mf:
                        cell.fill = mf
                if isinstance(cell.value, float):
                    cell.number_format = "0.0%" if str(h).endswith("%") else "0.000"

        # Master Matchups: make it readable and cap to top 100 visible rows
        if ws.title == "Master Matchups":
            ws.freeze_panes = "A2"
            ws.sheet_properties.tabColor = "D4A017"
            ws.insert_rows(1)
            ws["A1"] = "🏆 MASTER MATCHUPS — TOP 100 VIEW"
            ws["A1"].fill = fill(NAVY)
            ws["A1"].font = f(14,True,WHITE)
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=min(ws.max_column,12))
            for r in range(104, ws.max_row + 1):
                ws.row_dimensions[r].hidden = True

        # Game tabs: add simple top summary card once
        if ws.title not in ["Slate","🚨 Command Center","Category Explorer","Master Matchups"] and not str(ws["A1"].value).startswith("⚾ GAME SUMMARY"):
            game_name = str(ws["A1"].value or ws.title)
            ws.insert_rows(1,8)
            ws["A1"] = f"⚾ GAME SUMMARY — {game_name}"
            ws["A1"].fill = fill(NAVY)
            ws["A1"].font = f(14,True,WHITE)
            ws.merge_cells("A1:N1")

            blocks = [("🔥 Hits","Hits",1),("💥 Total Bases","Total Bases",6),("💣 Home Runs","Home Runs",11)]
            for label, market, start_col in blocks:
                ws.cell(3,start_col,label).fill = fill(BLUE)
                ws.cell(3,start_col).font = f(10,True,TEXT)
                ws.merge_cells(start_row=3,start_column=start_col,end_row=3,end_column=start_col+3)
                ws.cell(4,start_col,"Rank").font=f(8,True)
                ws.cell(4,start_col+1,"Player").font=f(8,True)
                ws.cell(4,start_col+2,"Score").font=f(8,True)
                ws.cell(4,start_col+3,"Risk").font=f(8,True)

                picks = top_for_game(game_name, market, 3)
                if not picks:
                    picks = top_for_game("", market, 3)
                for i,pick in enumerate(picks,1):
                    rr=4+i
                    vals=[i,pick.get("Hitter",""),pick.get("Score",""),pick.get("Risk","")]
                    for j,v in enumerate(vals):
                        cell=ws.cell(rr,start_col+j,v)
                        cell.fill=fill(LIGHT)
                        cell.border=border
                        cell.font=f(8, j==1)
                        if j==2:
                            cell.fill=metric_fill("Score",v) or fill(LIGHT)

            ws.freeze_panes = "A10"
            ws.sheet_properties.tabColor = NAVY

        # Bullpen status labels if bullpen table exists
        for r in range(1, min(ws.max_row,500)+1):
            vals = [ws.cell(r,c).value for c in range(1,min(ws.max_column,15)+1)]
            if "Reliever" in vals and "Pitches" in vals and "Status" not in vals:
                pitch_col = vals.index("Pitches") + 1
                status_col = pitch_col + 1
                ws.cell(r,status_col,"Status").fill = fill(BLUE)
                ws.cell(r,status_col).font = f(8,True,TEXT)
                for rr in range(r+1, min(r+25,ws.max_row)+1):
                    try: pitches = float(ws.cell(rr,pitch_col).value or 0)
                    except Exception: pitches = 0
                    if ws.cell(rr,1).value is None:
                        continue
                    status = "🔥 Burned" if pitches >= 25 else "⚠ Watch" if pitches >= 15 else "✅ Fresh"
                    cell=ws.cell(rr,status_col,status)
                    cell.fill = fill(RED if pitches >= 25 else YELLOW if pitches >= 15 else GREEN)
                    cell.font = f(8,True,TEXT)

        # Widths
        preferred = {
            "Rank":7,"Player":20,"Hitter":20,"Team":18,"Game":32,"Opp SP":18,"Best Market":16,
            "Pitch":8,"Side":8,"Usage":8,"Pitches":9,"PA":8,"Hits":8,"1B":8,"TB":8,
            "AVG":8,"SLG":8,"ISO":8,"wOBA":8,"xwOBA":8,"xBA":8,"xSLG":8,
            "SweetSpot%":11,"Barrel%":9,"HardHit%":10,"Whiff%":9,"Score":8,"Tier":7,
            "Why":44,"Risk":22,"Status":12
        }
        for c in range(1, min(ws.max_column,30)+1):
            header = None
            for r in range(1, min(ws.max_row,25)+1):
                v = ws.cell(r,c).value
                if str(v) in preferred:
                    header = str(v)
                    break
            ws.column_dimensions[get_column_letter(c)].width = preferred.get(header, 12)

    print("✅ final product polish applied: light theme, game summary cards, bullpen status, Master top 100")

def build_workbook(games, sc):
    wb = Workbook()
    wb.remove(wb.active)

    build_slate_sheet(wb, games)

    master_rows = []
    for i, g in enumerate(games, 1):
        away_hitters = team_hitter_pool(sc, g["away_code"])
        home_hitters = team_hitter_pool(sc, g["home_code"])
        away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
        home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

        home_vs_away = hitter_vs_pitch_rows(sc, home_hitters, away_mix)
        away_vs_home = hitter_vs_pitch_rows(sc, away_hitters, home_mix)

        for r in home_vs_away:
            master_rows.append({"Game #": i, "Game": g["game"], "Hitter Team": g["home_team"], "Opp SP": g["away_sp"], **r})
        for r in away_vs_home:
            master_rows.append({"Game #": i, "Game": g["game"], "Hitter Team": g["away_team"], "Opp SP": g["home_sp"], **r})

    master_rows.sort(key=lambda r: (r["Game #"], -r["wOBA"]))

    ws = wb.create_sheet("Master Matchups")
    headers = ["Game #", "Game", "Hitter Team", "Opp SP", "Pitch", "Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "xBA", "xSLG", "SweetSpot%", "Barrel%", "HardHit%", "Whiff%"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    color_cols = PCT_COLS | DEC_COLS
    for r, row in enumerate(master_rows, 2):
        band = GAME_FILL_A if row["Game #"] % 2 == 0 else GAME_FILL_B
        for c, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = band
            if h in PCT_COLS and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
            elif h in DEC_COLS and isinstance(val, (int, float)):
                
try:
    cell.value = float(cell.value)
except:
    pass
cell.number_format = "0.000"


    last_row = len(master_rows) + 1
    for c, h in enumerate(headers, 1):
        if h in color_cols and last_row > 1:
            apply_color_scale(ws, get_column_letter(c), 2, last_row, invert=(h in GOOD_LOW_STATS))

    widths = [7, 38, 22, 20, 7, 18, 6, 7, 7, 7, 7, 7, 8, 9, 8]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "E2"
    if master_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    for i, g in enumerate(games, 1):
        build_game_sheet(wb, i, g, sc)
    build_betting_command_center(wb, games, sc)
    mlb_lab_apply_visual_cohesion(wb)
    mlb_lab_final_product_polish(wb)
    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")

def main():
    games = get_schedule()
    sc = load_statcast()

    report = f"""# MLB-LAB V5.1 Pitch Matchup Engine

Date: {TODAY}

This runner uses:
- MLB Stats API for slate, teams, parks, probable pitchers
- Baseball Savant / Statcast for pitcher arsenal, L/R splits, hitter pools, hitter-vs-pitch matchups
- MLB Stats API boxscores for bullpen usage/fatigue over the last 4 days

Removed:
- FanGraphs hard dependency
- sportsbook odds
- CSV dependency
- scoring gimmicks
- weak-pitch labels
- manual placeholder tables

---

# Slate

| # | Game | Park | Away SP | Home SP |
|---:|---|---|---|---|
"""

    for i, g in enumerate(games, 1):
        report += f"| {i} | {g['game']} | {g['park']} | {g['away_sp']} | {g['home_sp']} |\n"

    report += "\n---\n\n# Full Game Breakdown Cards\n"

    for i, g in enumerate(games, 1):
        report += game_card(i, g, sc)

    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(report, encoding="utf-8")
    print(f"Updated {REPORT}")

    build_workbook(games, sc)

# ===== MLB-LAB EMERGENCY V5 ADD-ON: lineups / last 5 / bullpen / conclusions =====

def last_five_starts_table(sc, pitcher_id):
    if pitcher_id is None or sc is None or sc.empty:
        return md_table(["Date","Opponent","IP","H","ER","BB","K","HR"], [["—","—","—","—","—","—","—","—"]])
    d = sc[sc["pitcher"] == pitcher_id].copy()
    if d.empty or "game_date" not in d.columns:
        return md_table(["Date","Opponent","IP","H","ER","BB","K","HR"], [["—","—","—","—","—","—","—","—"]])
    games = []
    for date_val, x in d.groupby("game_date"):
        outs = x["outs_when_up"].max() if "outs_when_up" in x.columns else None
        ip = round((len(x[x["events"].notna()]) / 3), 1) if "events" in x.columns else "—"
        h = int(x["events"].isin(["single","double","triple","home_run"]).sum()) if "events" in x.columns else 0
        er = int(x["events"].isin(["home_run"]).sum()) if "events" in x.columns else 0
        bb = int(x["events"].isin(["walk","hit_by_pitch"]).sum()) if "events" in x.columns else 0
        k = int(x["events"].isin(["strikeout","strikeout_double_play"]).sum()) if "events" in x.columns else 0
        hr = int(x["events"].isin(["home_run"]).sum()) if "events" in x.columns else 0
        opp = x["home_team"].iloc[0] if "home_team" in x.columns else "—"
        games.append([str(date_val), opp, ip, h, er, bb, k, hr])
    games = sorted(games, key=lambda r: r[0], reverse=True)[:5]
    return md_table(["Date","Opponent","IP","H","ER","BB","K","HR"], games or [["—","—","—","—","—","—","—","—"]])


def projected_lineup_table(hitter_pool):
    if hitter_pool is None or hitter_pool.empty:
        return md_table(["Order","Hitter","PA","AVG","SLG","ISO","wOBA","Barrel%","HardHit%"], [["—","—","—","—","—","—","—","—","—"]])
    df = hitter_pool.copy()
    if "PA" in df.columns:
        df = df.sort_values("PA", ascending=False)
    rows = []
    for n, (_, r) in enumerate(df.head(9).iterrows(), 1):
        rows.append([
            n,
            r.get("Hitter", r.get("player_name", r.get("Name", "—"))),
            r.get("PA", "—"),
            fmt(r.get("AVG", "—")),
            fmt(r.get("SLG", "—")),
            fmt(r.get("ISO", "—")),
            fmt(r.get("wOBA", "—")),
            pct(r.get("Barrel%", r.get("barrel", "—"))),
            pct(r.get("HardHit%", r.get("hard_hit", "—"))),
        ])
    return md_table(["Order","Hitter","PA","AVG","SLG","ISO","wOBA","Barrel%","HardHit%"], rows)


def bullpen_context_table(sc, team_code):
    if sc is None or sc.empty or not team_code:
        return md_table(["Metric","Value"], [["Bullpen/Staff Sample","—"]])
    d = sc.copy()
    team_cols = [c for c in ["pitcher_team","home_team","away_team"] if c in d.columns]
    if team_cols:
        d = d[(d[team_cols[0]] == team_code) | (d.get("home_team", "") == team_code) | (d.get("away_team", "") == team_code)]
    events = d[d["events"].notna()] if "events" in d.columns else d
    if events.empty:
        return md_table(["Metric","Value"], [["Bullpen/Staff Sample","—"]])
    bf = len(events)
    h = int(events["events"].isin(["single","double","triple","home_run"]).sum())
    bb = int(events["events"].isin(["walk","hit_by_pitch"]).sum())
    k = int(events["events"].isin(["strikeout","strikeout_double_play"]).sum())
    hr = int(events["events"].isin(["home_run"]).sum())
    return md_table(["Metric","Value"], [
        ["Recent Staff BF", bf],
        ["Hits Allowed", h],
        ["Walks/HBP", bb],
        ["Strikeouts", k],
        ["Home Runs Allowed", hr],
        ["K Event Rate", pct(k / bf if bf else 0)],
        ["BB/HBP Event Rate", pct(bb / bf if bf else 0)],
        ["HR Event Rate", pct(hr / bf if bf else 0)],
    ])


def automated_conclusions(g, away_mix, home_mix, away_hitters, home_hitters):
    away_pitches = ", ".join(away_mix) if away_mix else "No current sample"
    home_pitches = ", ".join(home_mix) if home_mix else "No current sample"

    def top_hitter(df):
        if df is None or df.empty:
            return "No hitter-pool sample"
        sort_col = "wOBA" if "wOBA" in df.columns else ("ISO" if "ISO" in df.columns else None)
        if sort_col:
            row = df.sort_values(sort_col, ascending=False).iloc[0]
        else:
            row = df.iloc[0]
        return f"{row.get('Hitter', row.get('player_name', row.get('Name','—')))} — {sort_col or 'sample'} {fmt(row.get(sort_col, '—')) if sort_col else '—'}"

    return f"""
## Final Game Dissection

- Away pitcher pitch mix to inspect: {away_pitches}
- Home pitcher pitch mix to inspect: {home_pitches}
- Strongest home attack candidate: {top_hitter(home_hitters)}
- Strongest away attack candidate: {top_hitter(away_hitters)}
- Lineup advantage: compare projected top 9 hitter pools below.
- Handedness advantage: use full arsenal vs L/R tables above.
- Bullpen context: use recent staff pressure table below.
- Final MLB-LAB read: prioritize pitch mix + hitter pool + L/R damage + current form.
"""

if __name__ == "__main__":
    main()
