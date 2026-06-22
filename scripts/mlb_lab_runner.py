from datetime import date, timedelta
from pathlib import Path
import requests
import pandas as pd
from pybaseball import statcast, playerid_reverse_lookup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

TODAY = date.today()
START = (TODAY - timedelta(days=120)).isoformat()
END = TODAY.isoformat()

REPORT = Path("reports/mlb-lab-v5-matchup-engine.md")
WORKBOOK = Path("reports/mlb-lab-v5-matchup-engine.xlsx")
CACHE = Path("data/cache/statcast_cache.pkl")

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri")
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FILL = PatternFill("solid", start_color="2E5395", end_color="2E5395")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SUBHEAD_FONT = Font(name="Calibri", bold=True, size=10, color="333333")
GAME_FILL_A = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
GAME_FILL_B = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

# Stats where HIGH = favorable for hitter (green high, red low)
GOOD_HIGH_STATS = {"AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%"}
# Stats where HIGH = favorable for pitcher / bad for hitter (red high, green low)
GOOD_LOW_STATS = {"Whiff%", "K%", "K Event Rate", "HR Event Rate"}

TEAM_CODES = {
    "Arizona Diamondbacks": "AZ", "Athletics": "ATH", "Atlanta Braves": "ATL",
    "Baltimore Orioles": "BAL", "Boston Red Sox": "BOS", "Chicago Cubs": "CHC",
    "Chicago White Sox": "CWS", "Cincinnati Reds": "CIN", "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL", "Detroit Tigers": "DET", "Houston Astros": "HOU",
    "Kansas City Royals": "KC", "Los Angeles Angels": "LAA", "Los Angeles Dodgers": "LAD",
    "Miami Marlins": "MIA", "Milwaukee Brewers": "MIL", "Minnesota Twins": "MIN",
    "New York Mets": "NYM", "New York Yankees": "NYY", "Philadelphia Phillies": "PHI",
    "Pittsburgh Pirates": "PIT", "San Diego Padres": "SD", "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA", "St. Louis Cardinals": "STL", "Tampa Bay Rays": "TB",
    "Texas Rangers": "TEX", "Toronto Blue Jays": "TOR", "Washington Nationals": "WSH",
}

def mlb_json(url):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return r.json()

def fmt(x, n=3):
    try:
        if x is None or pd.isna(x):
            return "—"
        return f"{float(x):.{n}f}"
    except Exception:
        return "—"

def pct(x):
    try:
        if x is None or pd.isna(x):
            return "—"
        return f"{float(x) * 100:.1f}%"
    except Exception:
        return "—"

def md_table(headers, rows):
    out = "| " + " | ".join(headers) + " |\n"
    out += "| " + " | ".join(["---"] * len(headers)) + " |\n"
    if not rows:
        rows = [["No data"] + ["—"] * (len(headers) - 1)]
    for row in rows:
        out += "| " + " | ".join(str(x) for x in row) + " |\n"
    return out

def get_schedule():
    url = f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={TODAY}&hydrate=probablePitcher,venue,team"
    data = mlb_json(url)
    games = []

    for d in data.get("dates", []):
        for g in d.get("games", []):
            away = g["teams"]["away"]["team"]
            home = g["teams"]["home"]["team"]
            away_sp = g["teams"]["away"].get("probablePitcher", {})
            home_sp = g["teams"]["home"].get("probablePitcher", {})

            games.append({
                "game": f"{away['name']} @ {home['name']}",
                "away_team": away["name"],
                "home_team": home["name"],
                "away_team_id": away["id"],
                "home_team_id": home["id"],
                "away_code": TEAM_CODES.get(away["name"], ""),
                "home_code": TEAM_CODES.get(home["name"], ""),
                "park": g["venue"]["name"],
                "time": g.get("gameDate", "—"),
                "away_sp": away_sp.get("fullName", "TBD"),
                "home_sp": home_sp.get("fullName", "TBD"),
                "away_sp_id": away_sp.get("id"),
                "home_sp_id": home_sp.get("id"),
            })

    return games

def load_statcast():
    CACHE.parent.mkdir(parents=True, exist_ok=True)

    try:
        print(f"Pulling Statcast {START} to {END}")
        sc = statcast(start_dt=START, end_dt=END)
        if not sc.empty:
            sc.to_pickle(CACHE)
            return prep_statcast(sc)
    except Exception as e:
        print(f"Statcast pull failed: {e}")

    if CACHE.exists():
        print("Using cached Statcast.")
        return prep_statcast(pd.read_pickle(CACHE))

    return pd.DataFrame()

def prep_statcast(sc):
    sc = sc.copy()

    if "inning_topbot" in sc.columns:
        sc["batter_team"] = sc.apply(
            lambda r: r.get("away_team") if r.get("inning_topbot") == "Top" else r.get("home_team"),
            axis=1
        )

    if "batter" in sc.columns:
        try:
            ids = sorted(sc["batter"].dropna().astype(int).unique().tolist())
            lookup = playerid_reverse_lookup(ids, key_type="mlbam")
            lookup["batter_name"] = lookup["name_first"] + " " + lookup["name_last"]
            sc["batter_name"] = sc["batter"].map(dict(zip(lookup["key_mlbam"], lookup["batter_name"])))
        except Exception as e:
            print(f"Batter lookup failed: {e}")
            sc["batter_name"] = ""

    return sc

def recent_completed_games(team_id, days_back=4):
    if not team_id:
        return []

    start = (TODAY - timedelta(days=days_back)).isoformat()
    end = (TODAY - timedelta(days=1)).isoformat()
    url = (f"https://statsapi.mlb.com/api/v1/schedule?sportId=1&teamId={team_id}"
           f"&startDate={start}&endDate={end}&gameType=R")

    try:
        data = mlb_json(url)
    except Exception as e:
        print(f"Schedule lookup failed for team {team_id}: {e}")
        return []

    game_pks = []
    for d in data.get("dates", []):
        for g in d.get("games", []):
            if g.get("status", {}).get("abstractGameState") == "Final":
                game_pks.append((d.get("date"), g.get("gamePk")))

    return sorted(game_pks)

def bullpen_usage(team_id, days_back=4):
    games = recent_completed_games(team_id, days_back)

    if not games:
        return [], set()

    rows = []
    appearance_count = {}
    yesterday = (TODAY - timedelta(days=1)).isoformat()

    for game_date, game_pk in games:
        try:
            box = mlb_json(f"https://statsapi.mlb.com/api/v1/game/{game_pk}/boxscore")
        except Exception as e:
            print(f"Boxscore lookup failed for game {game_pk}: {e}")
            continue

        side = None
        for s in ("home", "away"):
            if box.get("teams", {}).get(s, {}).get("team", {}).get("id") == team_id:
                side = s
                break

        if not side:
            continue

        team_box = box["teams"][side]
        pitcher_ids = team_box.get("pitchers", [])
        players = team_box.get("players", {})

        # First pitcher listed for a team in a completed game is the starter;
        # everyone after is bullpen usage.
        for idx, pid in enumerate(pitcher_ids):
            stats = players.get(f"ID{pid}", {}).get("stats", {}).get("pitching", {})

            if not stats or idx == 0:
                continue

            name = players.get(f"ID{pid}", {}).get("person", {}).get("fullName", "Unknown")
            ip = stats.get("inningsPitched", "0.0")
            pitches = stats.get("numberOfPitches", 0)

            rows.append([name, game_date, ip, pitches])
            appearance_count[name] = appearance_count.get(name, 0) + 1

    tired = {name for name, count in appearance_count.items() if count >= 3}
    tired |= {r[0] for r in rows if r[1] == yesterday}

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows, tired

def bullpen_context(team_id, team_name, days_back=4):
    headers = ["Reliever", "Date", "IP", "Pitches"]
    rows, tired = bullpen_usage(team_id, days_back)

    flag = ", ".join(sorted(tired)) if tired else "No relievers flagged for heavy recent use"

    return f"""### {team_name} Bullpen — Last {days_back} Days

{md_table(headers, rows)}

**Caution arms (pitched yesterday, or 3+ appearances in window):** {flag}
"""

def pitcher_rows(sc, name, pitcher_id=None):
    if sc.empty or name == "TBD":
        return pd.DataFrame()

    if pitcher_id and "pitcher" in sc.columns:
        by_id = sc[sc["pitcher"] == pitcher_id]
        if not by_id.empty:
            return by_id

    if "player_name" not in sc.columns:
        return pd.DataFrame()

    parts = name.split()
    if len(parts) < 2:
        return pd.DataFrame()

    reverse = f"{parts[-1]}, {' '.join(parts[:-1])}"
    return sc[sc["player_name"].astype(str).str.lower() == reverse.lower()]

def event_stats(df):
    if df.empty or "events" not in df.columns:
        return 0, 0, 0, None, None, 0, 0, 0, 0

    events = df.dropna(subset=["events"])
    ab_events = events[~events["events"].isin(["walk", "intent_walk", "hit_by_pitch", "sac_bunt", "sac_fly"])]
    ab = len(ab_events)

    singles = (events["events"] == "single").sum()
    doubles = (events["events"] == "double").sum()
    triples = (events["events"] == "triple").sum()
    hrs = (events["events"] == "home_run").sum()

    hits = singles + doubles + triples + hrs
    tb = singles + 2 * doubles + 3 * triples + 4 * hrs

    avg = hits / ab if ab else 0
    slg = tb / ab if ab else 0
    iso = slg - avg if ab else 0

    woba = df["woba_value"].mean() if "woba_value" in df.columns else None
    xwoba = df["estimated_woba_using_speedangle"].mean() if "estimated_woba_using_speedangle" in df.columns else None

    bbe = df["launch_speed"].dropna() if "launch_speed" in df.columns else pd.Series(dtype=float)
    hard = (bbe >= 95).mean() if len(bbe) else 0

    barrel = 0
    if "launch_speed_angle" in df.columns:
        lsa = df["launch_speed_angle"].dropna()
        barrel = (lsa == 6).mean() if len(lsa) else 0

    swings = df[df["description"].isin(["swinging_strike", "swinging_strike_blocked", "foul", "foul_tip", "hit_into_play"])]
    whiffs = df[df["description"].isin(["swinging_strike", "swinging_strike_blocked"])]
    whiff = len(whiffs) / len(swings) if len(swings) else 0

    return avg, slg, iso, woba, xwoba, barrel, hard, whiff, len(events)

def pitcher_profile(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)

    if p.empty:
        return md_table(["Stat", "Value"], [])

    events = p.dropna(subset=["events"])
    h = events["events"].isin(["single", "double", "triple", "home_run"]).sum()
    bb = events["events"].isin(["walk", "intent_walk"]).sum()
    k = events["events"].isin(["strikeout", "strikeout_double_play"]).sum()
    hr = (events["events"] == "home_run").sum()

    total_events = max(1, len(events))

    rows = [
        ["Sample Pitches", len(p)],
        ["Batted/Result Events", len(events)],
        ["Hits Allowed", h],
        ["Walks", bb],
        ["Strikeouts", k],
        ["Home Runs", hr],
        ["K Event Rate", pct(k / total_events)],
        ["BB Event Rate", pct(bb / total_events)],
        ["HR Event Rate", pct(hr / total_events)],
    ]

    return md_table(["Stat", "Value"], rows)

def arsenal(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    headers = ["Pitch", "Side", "Usage", "Pitches", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]

    if p.empty or "pitch_type" not in p.columns or "stand" not in p.columns:
        return md_table(headers, [])

    rows = []
    total = max(1, len(p))

    for (pitch, side), g in p.groupby(["pitch_type", "stand"]):
        avg, slg, iso, woba, xwoba, barrel, hard, whiff, pa = event_stats(g)
        rows.append([
            pitch,
            f"vs {side}",
            pct(len(g) / total),
            len(g),
            fmt(avg),
            fmt(slg),
            fmt(iso),
            fmt(woba),
            fmt(xwoba),
            pct(barrel),
            pct(hard),
            pct(whiff),
        ])

    return md_table(headers, sorted(rows, key=lambda x: (x[0], x[1])))

def major_pitches(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    if p.empty or "pitch_type" not in p.columns:
        return []

    mix = p["pitch_type"].value_counts(normalize=True)
    return list(mix[mix >= 0.08].index)

def last_5_starts(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    headers = ["Date", "Pitches", "Hits", "BB", "K", "HR"]

    if p.empty or "game_date" not in p.columns:
        return md_table(headers, [])

    rows = []

    for game_date, g in p.groupby("game_date"):
        events = g.dropna(subset=["events"])
        hits = events["events"].isin(["single", "double", "triple", "home_run"]).sum()
        bb = events["events"].isin(["walk", "intent_walk"]).sum()
        k = events["events"].isin(["strikeout", "strikeout_double_play"]).sum()
        hr = (events["events"] == "home_run").sum()
        rows.append([game_date, len(g), hits, bb, k, hr])

    rows = sorted(rows, key=lambda x: str(x[0]), reverse=True)[:5]
    return md_table(headers, rows)

def team_hitter_pool(sc, team_code):
    if sc.empty or not team_code or "batter_team" not in sc.columns or "batter_name" not in sc.columns:
        return pd.DataFrame()

    t = sc[sc["batter_team"] == team_code]
    rows = []

    for name, g in t.groupby("batter_name"):
        if not name or pd.isna(name):
            continue

        avg, slg, iso, woba, xwoba, barrel, hard, whiff, pa = event_stats(g)

        if pa < 10:
            continue

        rows.append({
            "Hitter": str(name).title(),
            "PA": pa,
            "AVG": avg,
            "SLG": slg,
            "ISO": iso,
            "wOBA": woba if woba is not None else 0,
            "xwOBA": xwoba if xwoba is not None else 0,
            "Barrel%": barrel,
            "HardHit%": hard,
            "Whiff%": whiff,
        })

    df = pd.DataFrame(rows)

    if df.empty or "PA" not in df.columns:
        return pd.DataFrame()

    return df.sort_values(["PA", "wOBA"], ascending=[False, False]).head(12)

def hitter_pool_md(df):
    headers = ["Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]

    if df.empty:
        return md_table(headers, [])

    rows = []
    for _, r in df.iterrows():
        rows.append([
            r["Hitter"],
            int(r["PA"]),
            fmt(r["AVG"]),
            fmt(r["SLG"]),
            fmt(r["ISO"]),
            fmt(r["wOBA"]),
            fmt(r["xwOBA"]),
            pct(r["Barrel%"]),
            pct(r["HardHit%"]),
            pct(r["Whiff%"]),
        ])

    return md_table(headers, rows)

def hitter_vs_pitch_rows(sc, hitters, pitches):
    if sc.empty or hitters.empty or not pitches or "batter_name" not in sc.columns or "pitch_type" not in sc.columns:
        return []

    names = set(hitters["Hitter"].astype(str).str.lower())
    rows = []

    for pitch in pitches:
        pitch_df = sc[sc["pitch_type"] == pitch]

        for name in names:
            h = pitch_df[pitch_df["batter_name"].astype(str).str.lower() == name]

            if h.empty:
                continue

            avg, slg, iso, woba, xwoba, barrel, hard, whiff, pa = event_stats(h)

            if pa < 3:
                continue

            rows.append({
                "Pitch": pitch, "Hitter": name.title(), "PA": pa,
                "AVG": avg, "SLG": slg, "ISO": iso,
                "wOBA": woba if pd.notna(woba) else 0,
                "xwOBA": xwoba if pd.notna(xwoba) else 0,
                "Barrel%": barrel, "HardHit%": hard, "Whiff%": whiff,
            })

    rows.sort(key=lambda r: r["wOBA"], reverse=True)
    return rows[:30]

def hitter_vs_pitch(sc, hitters, pitches):
    headers = ["Pitch", "Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]

    if sc.empty or hitters.empty or not pitches or "batter_name" not in sc.columns or "pitch_type" not in sc.columns:
        return md_table(headers, [])

    names = set(hitters["Hitter"].astype(str).str.lower())
    rows = []

    for pitch in pitches:
        pitch_df = sc[sc["pitch_type"] == pitch]

        for name in names:
            h = pitch_df[pitch_df["batter_name"].astype(str).str.lower() == name]

            if h.empty:
                continue

            avg, slg, iso, woba, xwoba, barrel, hard, whiff, pa = event_stats(h)

            if pa < 3:
                continue

            rows.append([
                pitch,
                name.title(),
                pa,
                fmt(avg),
                fmt(slg),
                fmt(iso),
                fmt(woba),
                fmt(xwoba),
                pct(barrel),
                pct(hard),
                pct(whiff),
            ])

    if not rows:
        return md_table(headers, [])

    rows = sorted(rows, key=lambda r: float(r[5]) if r[5] != "—" else 0, reverse=True)[:30]
    return md_table(headers, rows)

def game_dissection(sc, g, away_hitters, home_hitters):
    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    return f"""
## Final Game Dissection

- Away pitcher pitch mix to inspect: {", ".join(away_mix) if away_mix else "No current Statcast sample"}
- Home pitcher pitch mix to inspect: {", ".join(home_mix) if home_mix else "No current Statcast sample"}
- Home hitters should be checked against away SP pitch mix above.
- Away hitters should be checked against home SP pitch mix above.
- Lineup advantage: compare wOBA, xwOBA, ISO, Barrel%, HardHit%, Whiff%.
- Handedness advantage: use pitcher arsenal vs L/R tables.
- Bullpen fatigue: see Bullpen Fatigue Report above.
- Final MLB-LAB read: decide from pitch mix, hitter pool, L/R damage, current form, and bullpen fatigue.
"""

def game_card(i, g, sc):
    away_hitters = team_hitter_pool(sc, g["away_code"])
    home_hitters = team_hitter_pool(sc, g["home_code"])

    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    return f"""
---

# {i}. {g['game']}

## Game Context

{md_table(["Field", "Value"], [
    ["Park", g["park"]],
    ["Time", g["time"]],
    ["Away Team", g["away_team"]],
    ["Home Team", g["home_team"]],
    ["Away Probable Pitcher", g["away_sp"]],
    ["Home Probable Pitcher", g["home_sp"]],
])}

## Away Starting Pitcher: {g['away_sp']}

### Pitcher Profile

{pitcher_profile(sc, g["away_sp"], g.get("away_sp_id"))}

### Full Pitch Arsenal vs L/R

{arsenal(sc, g["away_sp"], g.get("away_sp_id"))}

### Last 5 Starts / Appearances

{last_5_starts(sc, g["away_sp"], g.get("away_sp_id"))}

## Home Hitters vs Away SP Pitch Mix

{hitter_vs_pitch(sc, home_hitters, away_mix)}

## Home Starting Pitcher: {g['home_sp']}

### Pitcher Profile

{pitcher_profile(sc, g["home_sp"], g.get("home_sp_id"))}

### Full Pitch Arsenal vs L/R

{arsenal(sc, g["home_sp"], g.get("home_sp_id"))}

### Last 5 Starts / Appearances

{last_5_starts(sc, g["home_sp"], g.get("home_sp_id"))}

## Away Hitters vs Home SP Pitch Mix

{hitter_vs_pitch(sc, away_hitters, home_mix)}

## {g['away_team']} Team Hitter Pool

{hitter_pool_md(away_hitters)}

## {g['home_team']} Team Hitter Pool

{hitter_pool_md(home_hitters)}

## Bullpen Fatigue Report

{bullpen_context(g.get("away_team_id"), g["away_team"])}

{bullpen_context(g.get("home_team_id"), g["home_team"])}

{game_dissection(sc, g, away_hitters, home_hitters)}
"""

PCT_COLS = {"Barrel%", "HardHit%", "Whiff%", "K Event Rate", "BB Event Rate", "HR Event Rate", "Usage"}
DEC_COLS = {"AVG", "SLG", "ISO", "wOBA", "xwOBA"}

def apply_color_scale(ws, col_letter, first_row, last_row, invert=False):
    if last_row < first_row:
        return
    lo, hi = ("FFC7CE", "C6EFCE") if not invert else ("C6EFCE", "FFC7CE")
    rule = ColorScaleRule(
        start_type="min", start_color=lo,
        mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color=hi,
    )
    ws.conditional_formatting.add(f"{col_letter}{first_row}:{col_letter}{last_row}", rule)

def style_table(ws, start_row, start_col, headers, rows, pct_cols, dec_cols, color_cols=None):
    """Writes a header+rows table at (start_row, start_col) with banding,
    number formats, and per-column color scales. Returns the row after the table."""
    color_cols = color_cols if color_cols is not None else (PCT_COLS | DEC_COLS | {"K%", "BB%", "HR Event Rate"})

    for c, h in enumerate(headers):
        col = start_col + c
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows):
        excel_row = start_row + 1 + r
        for c, h in enumerate(headers):
            col = start_col + c
            val = row.get(h, "")
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if h in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
            elif h in dec_cols and isinstance(val, (int, float)):
                cell.number_format = "0.000"

    last_row = start_row + len(rows)
    if rows:
        for c, h in enumerate(headers):
            if h in color_cols:
                col_letter = get_column_letter(start_col + c)
                invert = h in GOOD_LOW_STATS
                apply_color_scale(ws, col_letter, start_row + 1, last_row, invert=invert)

    return last_row

def section_header(ws, row, col, text, span):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return row + 1

def build_slate_sheet(wb, games):
    ws = wb.create_sheet("Slate")
    headers = ["#", "Game", "Park", "Away SP", "Home SP", "Time (UTC)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, g in enumerate(games, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=g["game"])
        ws.cell(row=i + 1, column=3, value=g["park"])
        ws.cell(row=i + 1, column=4, value=g["away_sp"])
        ws.cell(row=i + 1, column=5, value=g["home_sp"])
        ws.cell(row=i + 1, column=6, value=g["time"])
        for c in range(1, 7):
            ws.cell(row=i + 1, column=c).font = BODY_FONT

    widths = [4, 40, 24, 22, 22, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(games) + 1}"
    return ws

def build_game_sheet(wb, i, g, sc):
    away_hitters = team_hitter_pool(sc, g["away_code"])
    home_hitters = team_hitter_pool(sc, g["home_code"])
    away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
    home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

    home_vs_away = hitter_vs_pitch_rows(sc, home_hitters, away_mix)
    away_vs_home = hitter_vs_pitch_rows(sc, away_hitters, home_mix)

    away_arsenal = arsenal_rows(sc, g["away_sp"], g.get("away_sp_id"))
    home_arsenal = arsenal_rows(sc, g["home_sp"], g.get("home_sp_id"))

    away_bullpen, away_tired = bullpen_usage(g.get("away_team_id"), 4)
    home_bullpen, home_tired = bullpen_usage(g.get("home_team_id"), 4)

    tab = f"G{i} {g['away_code']}@{g['home_code']}"
    ws = wb.create_sheet(tab[:31])

    ws["A1"] = g["game"]
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{g['park']}  |  {g['time']}  |  Away SP: {g['away_sp']}  |  Home SP: {g['home_sp']}"
    ws["A2"].font = SUBHEAD_FONT

    arsenal_headers = ["Pitch", "Side", "Usage", "Pitches", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]
    matchup_headers = ["Pitch", "Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]
    pool_headers = ["Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]
    bullpen_headers = ["Reliever", "Date", "IP", "Pitches"]

    row = 4
    # --- AWAY SP arsenal, then HOME hitters vs that arsenal (the actual matchup) ---
    row = section_header(ws, row, 1, f"{g['away_sp']} ({g['away_team']}) — Pitch Arsenal", len(arsenal_headers))
    row = style_table(ws, row, 1, arsenal_headers, away_arsenal, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['home_team']} Hitters vs {g['away_sp']} Pitch Mix", len(matchup_headers))
    row = style_table(ws, row, 1, matchup_headers, home_vs_away, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['home_team']} Hitter Pool (Season)", len(pool_headers))
    row = style_table(ws, row, 1, pool_headers, home_hitters.to_dict("records"), PCT_COLS, DEC_COLS) + 2

    # --- HOME SP arsenal, then AWAY hitters vs that arsenal ---
    row = section_header(ws, row, 1, f"{g['home_sp']} ({g['home_team']}) — Pitch Arsenal", len(arsenal_headers))
    row = style_table(ws, row, 1, arsenal_headers, home_arsenal, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['away_team']} Hitters vs {g['home_sp']} Pitch Mix", len(matchup_headers))
    row = style_table(ws, row, 1, matchup_headers, away_vs_home, PCT_COLS, DEC_COLS) + 2

    row = section_header(ws, row, 1, f"{g['away_team']} Hitter Pool (Season)", len(pool_headers))
    row = style_table(ws, row, 1, pool_headers, away_hitters.to_dict("records"), PCT_COLS, DEC_COLS) + 2

    # --- Bullpen fatigue, both teams ---
    row = section_header(ws, row, 1, f"{g['away_team']} Bullpen — Last 4 Days", len(bullpen_headers))
    away_bp_rows = [{"Reliever": r[0], "Date": r[1], "IP": r[2], "Pitches": r[3]} for r in away_bullpen]
    row = style_table(ws, row, 1, bullpen_headers, away_bp_rows, set(), set(), color_cols=set())
    if away_tired:
        ws.cell(row=row, column=1, value=f"Caution arms: {', '.join(sorted(away_tired))}").font = Font(italic=True, color="C00000")
    row += 2

    row = section_header(ws, row, 1, f"{g['home_team']} Bullpen — Last 4 Days", len(bullpen_headers))
    home_bp_rows = [{"Reliever": r[0], "Date": r[1], "IP": r[2], "Pitches": r[3]} for r in home_bullpen]
    row = style_table(ws, row, 1, bullpen_headers, home_bp_rows, set(), set(), color_cols=set())
    if home_tired:
        ws.cell(row=row, column=1, value=f"Caution arms: {', '.join(sorted(home_tired))}").font = Font(italic=True, color="C00000")

    widths = [18, 22, 8, 8, 8, 8, 8, 8, 9, 10, 9, 9]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws

def arsenal_rows(sc, name, pitcher_id=None):
    p = pitcher_rows(sc, name, pitcher_id)
    if p.empty or "pitch_type" not in p.columns or "stand" not in p.columns:
        return []

    rows = []
    total = max(1, len(p))
    for (pitch, side), g in p.groupby(["pitch_type", "stand"]):
        avg, slg, iso, woba, xwoba, barrel, hard, whiff, pa = event_stats(g)
        rows.append({
            "Pitch": pitch, "Side": f"vs {side}", "Usage": len(g) / total, "Pitches": len(g),
            "AVG": avg, "SLG": slg, "ISO": iso,
            "wOBA": woba if pd.notna(woba) else 0,
            "xwOBA": xwoba if pd.notna(xwoba) else 0,
            "Barrel%": barrel, "HardHit%": hard, "Whiff%": whiff,
        })
    return sorted(rows, key=lambda r: (r["Pitch"], r["Side"]))

def build_workbook(games, sc):
    wb = Workbook()
    wb.remove(wb.active)

    build_slate_sheet(wb, games)

    master_rows = []
    for i, g in enumerate(games, 1):
        away_hitters = team_hitter_pool(sc, g["away_code"])
        home_hitters = team_hitter_pool(sc, g["home_code"])
        away_mix = major_pitches(sc, g["away_sp"], g.get("away_sp_id"))
        home_mix = major_pitches(sc, g["home_sp"], g.get("home_sp_id"))

        home_vs_away = hitter_vs_pitch_rows(sc, home_hitters, away_mix)
        away_vs_home = hitter_vs_pitch_rows(sc, away_hitters, home_mix)

        for r in home_vs_away:
            master_rows.append({"Game #": i, "Game": g["game"], "Hitter Team": g["home_team"], "Opp SP": g["away_sp"], **r})
        for r in away_vs_home:
            master_rows.append({"Game #": i, "Game": g["game"], "Hitter Team": g["away_team"], "Opp SP": g["home_sp"], **r})

    master_rows.sort(key=lambda r: (r["Game #"], -r["wOBA"]))

    ws = wb.create_sheet("Master Matchups")
    headers = ["Game #", "Game", "Hitter Team", "Opp SP", "Pitch", "Hitter", "PA", "AVG", "SLG", "ISO", "wOBA", "xwOBA", "Barrel%", "HardHit%", "Whiff%"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    color_cols = PCT_COLS | DEC_COLS
    for r, row in enumerate(master_rows, 2):
        band = GAME_FILL_A if row["Game #"] % 2 == 0 else GAME_FILL_B
        for c, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = band
            if h in PCT_COLS and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
            elif h in DEC_COLS and isinstance(val, (int, float)):
                cell.number_format = "0.000"

    last_row = len(master_rows) + 1
    for c, h in enumerate(headers, 1):
        if h in color_cols and last_row > 1:
            apply_color_scale(ws, get_column_letter(c), 2, last_row, invert=(h in GOOD_LOW_STATS))

    widths = [7, 38, 22, 20, 7, 18, 6, 7, 7, 7, 7, 7, 8, 9, 8]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "E2"
    if master_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    for i, g in enumerate(games, 1):
        build_game_sheet(wb, i, g, sc)

    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")

def main():
    games = get_schedule()
    sc = load_statcast()

    report = f"""# MLB-LAB V5.1 Pitch Matchup Engine

Date: {TODAY}

This runner uses:
- MLB Stats API for slate, teams, parks, probable pitchers
- Baseball Savant / Statcast for pitcher arsenal, L/R splits, hitter pools, hitter-vs-pitch matchups
- MLB Stats API boxscores for bullpen usage/fatigue over the last 4 days

Removed:
- FanGraphs hard dependency
- sportsbook odds
- CSV dependency
- scoring gimmicks
- weak-pitch labels
- manual placeholder tables

---

# Slate

| # | Game | Park | Away SP | Home SP |
|---:|---|---|---|---|
"""

    for i, g in enumerate(games, 1):
        report += f"| {i} | {g['game']} | {g['park']} | {g['away_sp']} | {g['home_sp']} |\n"

    report += "\n---\n\n# Full Game Breakdown Cards\n"

    for i, g in enumerate(games, 1):
        report += game_card(i, g, sc)

    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(report, encoding="utf-8")
    print(f"Updated {REPORT}")

    build_workbook(games, sc)

if __name__ == "__main__":
    main()
